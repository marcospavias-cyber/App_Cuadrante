import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from datetime import timedelta
import io
import random
import calendar
from itertools import groupby
from operator import itemgetter

# ==============================================================================
# 1. CONFIGURACIÓN Y CONSTANTES
# ==============================================================================

st.set_page_config(layout="wide", page_title="Gestor V52.0 - Final Precision", page_icon="🚒")

# --- CRÉDITOS SUPERIORES ---
st.markdown("<h5 style='text-align: center; color: #888;'>Diseñado por Marcos Esteban Vives</h5>", unsafe_allow_html=True)
st.title("🚒 Gestor V52.0: Sistema Integral de Precisión")

# --- MANUAL Y AVISO ---
with st.expander("📘 MANUAL DE USUARIO (Haz clic para desplegar)", expanded=False):
    st.markdown("""
    ### 🎯 Precisión Total en Excel
    * **Cálculo Real:** La hoja de estadísticas ahora suma tus periodos fecha a fecha, ignorando si te ha tocado cubrir una guardia en medio. Tus 39 días son sagrados.
    * **Visualización:** Se distinguen claramente los días de guardia (Oro) de los libres (Crema).

    ### 🚀 Rendimiento
    * Si la aplicación va lenta tras muchos cambios, guarda tu Backup y **recarga la página (F5)** para limpiar la memoria.

    ### ⚙️ Flujo de Trabajo
    1. **Configura:** Estrategia y Nocturnas.
    2. **Rellena:** Usa bloques estándar o "A Medida".
    3. **Genera:** El botón "Automático" respeta lo que ya has puesto manual.
    """)

st.warning("⚠️ AVISO: El Excel final es la fuente de verdad. Revisa la hoja 'Estadísticas' para confirmar el cómputo anual.")

TEAMS = ['A', 'B', 'C']
ROLES = ["Jefe", "Subjefe", "Conductor", "Bombero"]
MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# --- ESTRATEGIAS ---
STRATEGIES = {
    "standard": {
        "name": "🛡️ Estándar Elástica (Recomendada)",
        "blocks": [
            {"dur": 10, "cred": 4, "label": "📦 Bloque 10d (4 Cr)"},
            {"dur": 10, "cred": 3, "label": "📦 Bloque 10d (3 Cr)"},
            {"dur": 9,  "cred": 3, "label": "📦 Bloque 9d (3 Cr)"},
            {"dur": 4,  "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1,  "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 10, "target": 4}, {"dur": 10, "target": 3}, 
            {"dur": 10, "target": 3}, {"dur": 9,  "target": 3}
        ]
    },
    "safe": {
        "name": "🔢 Matemática Pura",
        "blocks": [
            {"dur": 12, "cred": 4, "label": "📦 Largo 12d (4 Cr)"},
            {"dur": 9,  "cred": 3, "label": "📦 Medio 9d (3 Cr)"},
            {"dur": 6,  "cred": 2, "label": "📦 Corto 6d (2 Cr)"},
            {"dur": 4,  "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1,  "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 12, "target": 4}, {"dur": 12, "target": 4}, 
            {"dur": 9, "target": 3}, {"dur": 6, "target": 2}
        ]
    },
    "balanced": {
        "name": "⚖️ Tridente",
        "blocks": [
            {"dur": 13, "cred": 5, "label": "📦 Bloque 13d (5 Cr)"},
            {"dur": 13, "cred": 4, "label": "📦 Bloque 13d (4 Cr)"},
            {"dur": 4,  "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1,  "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 13, "target": 5}, {"dur": 13, "target": 4}, {"dur": 13, "target": 4}
        ]
    },
    "long": {
        "name": "✈️ Larga Estancia",
        "blocks": [
            {"dur": 15, "cred": 5, "label": "📦 Gran Viaje 15d (5 Cr)"},
            {"dur": 9,  "cred": 3, "label": "📦 Escapada 9d (3 Cr)"},
            {"dur": 4,  "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1,  "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 15, "target": 5}, {"dur": 15, "target": 5}, {"dur": 9, "target": 3}
        ]
    },
    "micro": {
        "name": "🐜 Hormiga",
        "blocks": [
            {"dur": 6, "cred": 2, "label": "📦 Semana 6d (2 Cr)"},
            {"dur": 9, "cred": 3, "label": "📦 Semana+ 9d (3 Cr)"},
            {"dur": 4, "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1, "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 6, "target": 2}, {"dur": 6, "target": 2}, 
            {"dur": 6, "target": 2}, {"dur": 6, "target": 2}, 
            {"dur": 6, "target": 2}, {"dur": 9, "target": 3}
        ]
    },
    "sniper": {
        "name": "🎯 Francotirador",
        "blocks": [
            {"dur": 1, "cred": 1, "label": "📦 Día Suelto (1 Cr)"},
            {"dur": 4, "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"} 
        ],
        "auto_recipe": [{"dur": 1, "target": 1}] * 13
    },
    "balanced_plus": {
        "name": "🧩 Flexible",
        "blocks": [
            {"dur": 8, "cred": 3, "label": "📦 8d (3 Cr)"},
            {"dur": 8, "cred": 2, "label": "📦 8d (2 Cr)"},
            {"dur": 7, "cred": 2, "label": "📦 7d (2 Cr)"},
            {"dur": 4, "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1, "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 8, "target": 3}, {"dur": 8, "target": 3}, {"dur": 8, "target": 3},
            {"dur": 8, "target": 2}, {"dur": 7, "target": 2}
        ]
    }
}

DEFAULT_ROSTER = [
    {"ID_Puesto": "Jefe A",       "Nombre": "Jefe A",       "Turno": "A", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe A",    "Nombre": "Subjefe A",    "Turno": "A", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond A",       "Nombre": "Cond A",       "Turno": "A", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero A1",   "Nombre": "Bombero A1",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A2",   "Nombre": "Bombero A2",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A3",   "Nombre": "Bombero A3",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Jefe B",       "Nombre": "Jefe B",       "Turno": "B", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe B",    "Nombre": "Subjefe B",    "Turno": "B", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond B",       "Nombre": "Cond B",       "Turno": "B", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero B1",   "Nombre": "Bombero B1",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B2",   "Nombre": "Bombero B2",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B3",   "Nombre": "Bombero B3",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Jefe C",       "Nombre": "Jefe C",       "Turno": "C", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe C",    "Nombre": "Subjefe C",    "Turno": "C", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond C",       "Nombre": "Cond C",       "Turno": "C", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero C1",   "Nombre": "Bombero C1",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C2",   "Nombre": "Bombero C2",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C3",   "Nombre": "Bombero C3",   "Turno": "C", "Rol": "Bombero",    "SV": False},
]

# ==============================================================================
# 2. LÓGICA DE NEGOCIO (OPTIMIZADA)
# ==============================================================================

def get_short_id(name, role, turn):
    if role == "Jefe": return f"J{turn}"
    if role == "Subjefe": return f"S{turn}"
    if role == "Conductor": return f"C{turn}"
    if "Bombero" in name:
        parts = name.split()
        if len(parts) > 1:
            suffix = parts[-1]
            if len(suffix) >= 2:
                return f"B{suffix[-1]}{turn}"
    return f"{name[:3]}{turn}"

@st.cache_data
def generate_base_schedule(year):
    is_leap = calendar.isleap(year)
    total_days = 366 if is_leap else 365
    status = {'A': 0, 'B': 2, 'C': 1} 
    schedule = {team: [] for team in TEAMS}
    for _ in range(total_days):
        for t in TEAMS:
            schedule[t].append('T' if status[t] == 0 else 'L')
            status[t] = (status[t] + 1) % 3
    return schedule, total_days

def is_in_night_period(day_idx, year, night_periods):
    current_date = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
    for start, end in night_periods:
        if start <= current_date <= end: return True
    return False

@st.cache_data
def get_night_transition_dates(night_periods):
    dates = set()
    for start, end in night_periods:
        dates.add(end) 
    return dates

def calculate_stats(roster_df, requests, year):
    base_sch, _ = generate_base_schedule(year)
    stats = {}
    for _, p in roster_df.iterrows():
        stats[p['Nombre']] = {'credits': 0, 'natural': 0}
    for req in requests:
        name = req['Nombre']
        if name not in stats: continue
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        row = roster_df[roster_df['Nombre'] == name].iloc[0]
        nat = (e_idx - s_idx) + 1
        cred = 0
        for d in range(s_idx, e_idx + 1):
            if base_sch[row['Turno']][d] == 'T': cred += 1
        stats[name]['credits'] += cred
        stats[name]['natural'] += nat
    return stats

def get_clustered_dates(available_idxs, needed_count):
    if not available_idxs: return []
    groups = []
    for k, g in groupby(enumerate(available_idxs), lambda ix: ix[0] - ix[1]):
        groups.append(list(map(itemgetter(1), g)))
    groups.sort(key=len, reverse=True)
    selected = []
    for group in groups:
        if len(selected) < needed_count:
            take = min(len(group), needed_count - len(selected))
            selected.extend(group[:take])
        else: break
    return sorted(selected)

# --- MAPA DE OCUPACIÓN ---
def precompute_occupancy(current_requests, roster_df, total_days):
    daily_absent = {i: [] for i in range(total_days)}
    daily_roles = {i: [] for i in range(total_days)}
    
    for req in current_requests:
        p_row = roster_df[roster_df['Nombre'] == req['Nombre']]
        if p_row.empty: continue
        p_row = p_row.iloc[0]
        
        s = req['Inicio'].timetuple().tm_yday - 1
        e = req['Fin'].timetuple().tm_yday - 1
        s = max(0, s); e = min(total_days - 1, e)
        
        for d in range(s, e + 1):
            daily_absent[d].append(req['Nombre'])
            daily_roles[d].append(p_row['Rol'])
            
    return daily_absent, daily_roles

# --- DETECTOR DE CONFLICTOS ---
def analyze_slot(start_idx, duration, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles):
    total_days = len(base_sch['A'])
    if start_idx + duration > total_days: return False, "Fuera de año", None
    
    my_start_natural = start_idx
    my_end_natural = start_idx + duration - 1

    # 1. Reglas Diarias
    for i in range(start_idx, start_idx + duration):
        d_obj = datetime.date(year, 1, 1) + timedelta(days=i)
        if d_obj in transition_dates:
            if base_sch[person['Turno']][i] == 'T': 
                return False, "Conflicto Nocturna (Día Transición)", "Nocturna"
            
        occupants_T = occupation_map_T.get(i, [])
        for occ in occupants_T:
            if occ['Turno'] == person['Turno']: 
                return False, f"Coincide turno T con {occ['Nombre']}", occ['Nombre']
    
    # 2. Capacidad Global
    for d_check in range(my_start_natural, my_end_natural + 1):
        absent_list = [x for x in daily_absent[d_check] if x != person['Nombre']]
        if len(absent_list) >= 2: 
            culprit = ", ".join(absent_list[:2])
            return False, f"Cupo lleno (2 pers) el día {d_check+1}", culprit

    # 3. Categoría
    if person['Rol'] != 'Bombero':
        for d_check in range(my_start_natural, my_end_natural + 1):
            people_today = daily_absent[d_check]
            roles_today = daily_roles[d_check]
            for p_name, p_role in zip(people_today, roles_today):
                if p_name != person['Nombre'] and p_role == person['Rol']:
                    return False, f"Coincidencia Categoría con {p_name}", p_name
                        
    return True, "OK", None

def book_slot_gen(start_idx, duration, person, occupation_map):
    for i in range(start_idx, start_idx + duration):
        if i not in occupation_map: occupation_map[i] = []
        occupation_map[i].append(person)

def get_available_blocks_for_person(person_name, roster_df, current_requests, year, night_periods, month_range, strategy_key):
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    person = roster_df[roster_df['Nombre'] == person_name].iloc[0]
    start_month_idx = MESES.index(month_range[0]) + 1
    end_month_idx = MESES.index(month_range[1]) + 1
    
    occupation_map_T = {i:[] for i in range(total_days)}
    my_current_slots = [] 
    for req in current_requests:
        p_req = roster_df[roster_df['Nombre'] == req['Nombre']].iloc[0]
        s = req['Inicio'].timetuple().tm_yday - 1
        e = req['Fin'].timetuple().tm_yday - 1
        if req['Nombre'] != person_name:
            for d in range(s, e+1):
                if base_sch[p_req['Turno']][d] == 'T': occupation_map_T[d].append(p_req)
        else:
            my_current_slots.append((s, e))

    daily_absent, daily_roles = precompute_occupancy(current_requests, roster_df, total_days)

    block_defs = STRATEGIES[strategy_key]['blocks']
    options = {b['label']: [] for b in block_defs}
    conflicts_log = {b['label']: [] for b in block_defs}
    
    # Pre-cálculo para combos
    valid_fragments = []

    for d in range(total_days): 
        d_date = datetime.date(year, 1, 1) + timedelta(days=d)
        if d_date.month < start_month_idx or d_date.month > end_month_idx: continue
        
        # Fragmentos para combos
        for frag_dur in range(3, 8):
            if d + frag_dur <= total_days:
                is_valid_frag, _, _ = analyze_slot(d, frag_dur, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles)
                if is_valid_frag:
                    overlap = any(not (d + frag_dur - 1 < ms[0] or d > ms[1]) for ms in my_current_slots)
                    if not overlap:
                        frag_cred = sum([1 for k in range(d, d+frag_dur) if base_sch[person['Turno']][k] == 'T'])
                        valid_fragments.append({
                            'start_idx': d, 'end_idx': d+frag_dur-1, 
                            'dur': frag_dur, 'cred': frag_cred, 
                            'date_obj': d_date
                        })

        for b_def in block_defs:
            duration = b_def['dur']
            target_cred = b_def['cred']
            label_key = b_def['label']
            
            if d + duration > total_days: continue

            # ANÁLISIS DEL SLOT
            is_valid, reason, culprit = analyze_slot(d, duration, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles)

            if is_valid:
                overlap = False
                for ms in my_current_slots:
                    if not (d + duration - 1 < ms[0] or d > ms[1]): overlap = True; break
                
                if not overlap:
                    credits = 0
                    for k in range(d, d+duration):
                        if base_sch[person['Turno']][k] == 'T': credits += 1
                    
                    if credits == target_cred:
                        start_date = d_date
                        end_date = start_date + timedelta(days=duration-1)
                        txt = f"{start_date.strftime('%d/%m')} - {end_date.strftime('%d/%m')}"
                        
                        score = 0
                        if start_date.day in [1, 11, 21]: score += 2
                        for ms in my_current_slots:
                            if abs(d - ms[1]) == 1 or abs((d + duration) - ms[0]) == 1:
                                score += 3
                        
                        icon = "⭐" if score >= 2 else ""
                        options[label_key].append({
                            'label': f"{icon} {txt}", 
                            'start': start_date, 'end': end_date, 'score': score, 'type': 'single'
                        })
            else:
                credits = 0
                for k in range(d, d+duration):
                    if base_sch[person['Turno']][k] == 'T': credits += 1
                
                if credits == target_cred:
                     conflicts_log[label_key].append({'start': d_date, 'reason': reason, 'culprit': culprit})

    unique_frags = []
    seen = set()
    for f in valid_fragments:
        if f['start_idx'] not in seen:
            unique_frags.append(f); seen.add(f['start_idx'])
            
    for b_def in block_defs:
        target_dur = b_def['dur']
        target_cred = b_def['cred']
        label_key = b_def['label']
        
        if target_dur >= 9:
            combos_found = 0
            for i in range(len(unique_frags)):
                if combos_found > 15: break 
                frag_A = unique_frags[i]
                
                for j in range(i + 1, len(unique_frags)):
                    frag_B = unique_frags[j]
                    if frag_B['start_idx'] <= frag_A['end_idx']: continue
                    
                    is_dur_ok = abs((frag_A['dur'] + frag_B['dur']) - target_dur) <= 1
                    is_cred_ok = (frag_A['cred'] + frag_B['cred']) == target_cred
                    
                    if is_dur_ok and is_cred_ok:
                        s1 = frag_A['date_obj'].strftime('%d/%m'); e1 = (frag_A['date_obj'] + timedelta(days=frag_A['dur']-1)).strftime('%d/%m')
                        s2 = frag_B['date_obj'].strftime('%d/%m'); e2 = (frag_B['date_obj'] + timedelta(days=frag_B['dur']-1)).strftime('%d/%m')
                        
                        options[label_key].append({
                            'label': f"✂️ Combo: {s1}-{e1} & {s2}-{e2}",
                            'parts': [frag_A, frag_B],
                            'score': -1,
                            'type': 'combo'
                        })
                        combos_found += 1
                        break 

    for k in options:
        options[k].sort(key=lambda x: x['score'], reverse=True)
        
    return options, conflicts_log, occupation_map_T, daily_absent, daily_roles

def auto_generate_schedule(roster_df, year, night_periods, strategy_key, current_reqs):
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    
    occupation_map_T = {i:[] for i in range(total_days)}
    daily_absent = {i: [] for i in range(total_days)}
    daily_roles = {i: [] for i in range(total_days)}
    
    # Cargar estado actual para el generador
    for req in current_reqs:
        p_row = roster_df[roster_df['Nombre'] == req['Nombre']]
        if p_row.empty: continue
        p_row = p_row.iloc[0]
        s = req['Inicio'].timetuple().tm_yday - 1
        e = req['Fin'].timetuple().tm_yday - 1
        s = max(0, s); e = min(total_days - 1, e)
        for d in range(s, e + 1):
            daily_absent[d].append(req['Nombre'])
            daily_roles[d].append(p_row['Rol'])
            if base_sch[p_row['Turno']][d] == 'T':
                occupation_map_T[d].append(p_row)

    generated_requests = []
    people = roster_df.to_dict('records')
    priority_order = ["Jefe", "Subjefe", "Conductor", "Bombero"]
    people.sort(key=lambda x: priority_order.index(x['Rol']))
    RECIPE = STRATEGIES[strategy_key]['auto_recipe']
    
    for person in people:
        my_slots = []
        credits_got = 0
        
        my_existing = [r for r in current_reqs if r['Nombre'] == person['Nombre']]
        for r in my_existing:
            s = r['Inicio'].timetuple().tm_yday - 1
            e = r['Fin'].timetuple().tm_yday - 1
            my_slots.append((s, (e-s)+1))
            for k in range(s, e+1):
                if 0 <= k < total_days and base_sch[person['Turno']][k] == 'T':
                    credits_got += 1
        
        if credits_got >= 13: continue

        current_recipe = RECIPE.copy()
        current_recipe.sort(key=lambda x: x['dur'], reverse=True)
        
        for block in current_recipe:
            if credits_got >= 13: break
            duration = block['dur']
            target = block['target']
            options = []
            
            for d in range(0, total_days - duration):
                if len(daily_absent[d]) >= 2: continue 
                c = sum([1 for k in range(d, d+duration) if base_sch[person['Turno']][k] == 'T'])
                if c == target:
                     is_valid, _, _ = analyze_slot(d, duration, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles)
                     if is_valid:
                         options.append(d)
            
            random.shuffle(options)
            
            for start in options:
                overlap = False
                for ms in my_slots:
                    if not (start + duration - 1 < ms[0] or start > ms[0] + ms[1] - 1): 
                        overlap = True; break
                
                if not overlap:
                    book_slot_gen(start, duration, person, occupation_map_T)
                    for k in range(start, start + duration):
                        daily_absent[k].append(person['Nombre'])
                        daily_roles[k].append(person['Rol'])
                    my_slots.append((start, duration))
                    credits_got += target
                    generated_requests.append({
                        "Nombre": person['Nombre'],
                        "Inicio": datetime.date(year, 1, 1) + timedelta(days=start),
                        "Fin": datetime.date(year, 1, 1) + timedelta(days=start+duration-1)
                    })
                    break 
        
        if credits_got < 13:
            all_days_random = list(range(total_days))
            random.shuffle(all_days_random)
            for d in all_days_random:
                if credits_got >= 13: break
                if base_sch[person['Turno']][d] == 'T':
                    is_valid, _, _ = analyze_slot(d, 1, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles)
                    if is_valid:
                        overlap = False
                        for ms in my_slots:
                            if not (d < ms[0] or d > ms[0] + ms[1] - 1): overlap = True; break
                        if not overlap:
                            book_slot_gen(d, 1, person, occupation_map_T)
                            for k in range(d, d+1):
                                daily_absent[k].append(person['Nombre'])
                                daily_roles[k].append(person['Rol'])
                            my_slots.append((d, 1))
                            credits_got += 1
                            generated_requests.append({
                                "Nombre": person['Nombre'],
                                "Inicio": datetime.date(year, 1, 1) + timedelta(days=d),
                                "Fin": datetime.date(year, 1, 1) + timedelta(days=d)
                            })
    return generated_requests

# --- RENDERIZADO VISUAL ---
@st.cache_data
def render_global_occupation_calendar(year, roster_df, requests, night_periods):
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    occ_map = {d: [] for d in range(total_days)}
    
    for req in requests:
        name = req['Nombre']
        if name not in roster_df['Nombre'].values: continue
        person_row = roster_df[roster_df['Nombre'] == name].iloc[0]
        turn = person_row['Turno']
        s = req['Inicio'].timetuple().tm_yday - 1
        e = req['Fin'].timetuple().tm_yday - 1
        for d in range(s, e+1):
            occ_map[d].append(get_short_id(name, person_row['Rol'], turn))

    html = "<div style='font-family:monospace; font-size:9px;'>"
    html += """
    <div style='display:flex; gap:10px; margin-bottom:10px; font-size:11px; font-weight:bold;'>
        <span style='background:#d4edda; color:#155724; padding:2px 6px; border:1px solid #c3e6cb;'>🟩 DISPONIBLE</span>
        <span style='background:#FFF3CD; color:#856404; padding:2px 6px; border:1px solid #FFEEBA;'>🟧 ÚLTIMA PLAZA</span>
        <span style='background:#F8D7DA; color:#721c24; padding:2px 6px; border:1px solid #F5C6CB;'>🟥 COMPLETO</span>
    </div>
    """
    html += "<div style='display:flex; margin-bottom:2px;'><div style='width:35px;'></div>"
    for d in range(1, 32):
        html += f"<div style='width:32px; text-align:center; color:#888;'>{d}</div>"
    html += "</div>"

    for m_idx, mes in enumerate(MESES):
        m_num = m_idx + 1
        days_in_month = calendar.monthrange(year, m_num)[1]
        html += f"<div style='display:flex; margin-bottom:2px;'><div style='width:35px; font-weight:bold; padding-top:8px;'>{mes}</div>"
        for d in range(1, 32):
            if d <= days_in_month:
                dt = datetime.date(year, m_num, d)
                d_idx = dt.timetuple().tm_yday - 1
                occupants = occ_map[d_idx]
                count = len(occupants)
                
                if count == 0: bg = "#d4edda"; txt_col = "#155724"
                elif count == 1: bg = "#FFF3CD"; txt_col = "#856404"
                else: bg = "#F8D7DA"; txt_col = "#721c24"

                border = "1px solid #fff"
                if dt in transition_dates: border = "2px solid red"
                label = "<br>".join(occupants)
                html += f"<div style='width:32px; height:30px; background-color:{bg}; color:{txt_col}; text-align:center; border:{border}; border-radius:2px; font-size:8px; line-height:9px; display:flex; align-items:center; justify-content:center;'>{label}</div>"
            else:
                html += "<div style='width:32px;'></div>"
        html += "</div>"
    html += "</div>"
    return html

def render_annual_calendar(year, team, base_sch, night_periods, custom_schedule=None):
    html = f"<div style='font-family:monospace; font-size:10px;'>"
    html += """
    <div style='display:flex; gap:10px; margin-bottom:5px; font-size:11px; font-weight:bold;'>
        <span style='background:#d4edda; color:#155724; padding:2px 5px; border:1px solid #c3e6cb;'>T (Guardia)</span>
        <span style='background:#FFC000; color:#000; padding:2px 5px; border:1px solid #DAA520;'>V (Pedido)</span>
        <span style='background:#FFFFE0; color:#555; padding:2px 5px; border:1px solid #EEE8AA;'>V(R) (Relleno)</span>
        <span style='background:#1E7E34; color:white; padding:2px 5px;'>T (Noche)</span>
        <span style='border:2px solid red; padding:0px 5px; color:red;'>Fin Noche</span>
    </div>
    """
    html += "<div style='display:flex; margin-bottom:2px;'><div style='width:30px;'></div>"
    for d in range(1, 32):
        html += f"<div style='width:20px; text-align:center; color:#888;'>{d}</div>"
    html += "</div>"
    for m_idx, mes in enumerate(MESES):
        m_num = m_idx + 1
        days_in_month = calendar.monthrange(year, m_num)[1]
        html += f"<div style='display:flex; margin-bottom:2px;'><div style='width:30px; font-weight:bold;'>{mes}</div>"
        for d in range(1, 32):
            if d <= days_in_month:
                dt = datetime.date(year, m_num, d)
                d_idx = dt.timetuple().tm_yday - 1
                state = base_sch[team][d_idx]
                final_val = state
                if custom_schedule: final_val = custom_schedule[d_idx]
                bg_color = "#eee"; text_color = "#ccc"; border = "1px solid #fff"
                if final_val == 'T': 
                    bg_color = "#d4edda"; text_color = "#155724"
                    if is_in_night_period(d_idx, year, night_periods):
                        bg_color = "#1E7E34"; text_color = "white"
                elif final_val == 'V':
                    bg_color = "#FFC000"; text_color = "#000"
                elif final_val == 'V(R)':
                    bg_color = "#FFFFE0"; text_color = "#555"
                elif final_val == 'T+':
                    bg_color = "#ADD8E6"; text_color = "#000"
                elif final_val == 'L*':
                    bg_color = "#E6E6FA"; text_color = "#000"
                if dt in get_night_transition_dates(night_periods): border = "2px solid red"
                html += f"<div style='width:20px; background-color:{bg_color}; color:{text_color}; text-align:center; border:{border}; border-radius:2px;'>{state[0]}</div>"
            else:
                html += "<div style='width:20px;'></div>"
        html += "</div>"
    html += "</div>"
    return html

# --- GESTIÓN DE COBERTURAS ---
def get_candidates(person_missing, roster_df, day_idx, current_schedule, year, night_periods, unavailable_map, adjustments_log_current_day=None):
    candidates = []
    missing_role = person_missing['Rol']
    missing_turn = person_missing['Turno']
    blocked_turns = set()
    
    if adjustments_log_current_day:
        for coverer_name in adjustments_log_current_day:
            cov_p = roster_df[roster_df['Nombre'] == coverer_name]
            if not cov_p.empty: blocked_turns.add(cov_p.iloc[0]['Turno'])
            
    turn_exhausted_from_night = None
    if day_idx > 0:
        prev_day_idx = day_idx - 1
        if is_in_night_period(prev_day_idx, year, night_periods):
            base_sch_temp, _ = generate_base_schedule(year)
            for t in TEAMS:
                if base_sch_temp[t][prev_day_idx] == 'T':
                    turn_exhausted_from_night = t; break
                    
    for _, candidate in roster_df.iterrows():
        cand_name = candidate['Nombre']
        
        if candidate['Turno'] == missing_turn: continue
        if cand_name in unavailable_map[day_idx]: continue
        
        cand_status = current_schedule[cand_name][day_idx]
        if cand_status != 'L': continue 
        
        if candidate['Turno'] in blocked_turns: continue
        if turn_exhausted_from_night and candidate['Turno'] == turn_exhausted_from_night: continue
        
        is_compatible = False
        cand_role = candidate['Rol']
        if missing_role == "Jefe" and cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Subjefe" and cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Conductor" and (cand_role == "Conductor" or candidate['SV']): is_compatible = True
        elif missing_role == "Bombero" and (cand_role == "Bombero" or candidate['SV']): is_compatible = True
        
        if is_compatible: candidates.append(cand_name)
        
    return candidates

def validate_and_generate_final(roster_df, requests, year, night_periods, forced_adjustments=None, strategy_key="standard"):
    if forced_adjustments is None: forced_adjustments = []
    base_schedule_turn, total_days = generate_base_schedule(year)
    final_schedule = {} 
    turn_coverage_counters = {'A': 0, 'B': 0, 'C': 0}
    person_coverage_counters = {name: 0 for name in roster_df['Nombre']}
    name_to_turn = {row['Nombre']: row['Turno'] for _, row in roster_df.iterrows()}
    
    for _, row in roster_df.iterrows():
        final_schedule[row['Nombre']] = base_schedule_turn[row['Turno']].copy()

    day_vacations = {i: [] for i in range(total_days)}
    natural_days_count = {name: 0 for name in roster_df['Nombre']}
    
    unavailable_map = {i: set() for i in range(total_days)}

    for req in requests:
        name = req['Nombre']
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        duration = (e_idx - s_idx) + 1
        natural_days_count[name] += duration
        
        for d in range(s_idx, e_idx + 1):
            unavailable_map[d].add(name)
            
            if final_schedule[name][d] == 'T':
                day_vacations[d].append(name)
                final_schedule[name][d] = 'V'
            else:
                final_schedule[name][d] = 'V(L)'

    adjustments_log = []
    
    for d in range(total_days):
        absent_people = day_vacations[d]
        if not absent_people: continue
        
        current_day_coverers = []
        absent_people.sort(key=lambda x: 0 if "Jefe" in x or "Subjefe" in x else 1)

        for name_missing in absent_people:
            person_row = roster_df[roster_df['Nombre'] == name_missing].iloc[0]
            candidates = get_candidates(
                person_row, roster_df, d, final_schedule, year, 
                night_periods, unavailable_map, current_day_coverers
            )
            
            if candidates:
                valid = []
                for c in candidates:
                    prev = final_schedule[c][d-1] if d > 0 else 'L'
                    prev2 = final_schedule[c][d-2] if d > 1 else 'L'
                    worked_prev = prev.startswith('T')
                    worked_prev2 = prev2.startswith('T')
                    if not (worked_prev and worked_prev2): 
                        valid.append(c)
                        
                if valid:
                    valid.sort(key=lambda x: (turn_coverage_counters[name_to_turn[x]], person_coverage_counters[x], random.random()))
                    chosen = valid[0]
                    final_schedule[chosen][d] = f"T*({name_missing})"
                    adjustments_log.append((d, chosen, name_missing))
                    current_day_coverers.append(chosen)
                    unavailable_map[d].add(chosen)
                    turn_coverage_counters[name_to_turn[chosen]] += 1
                    person_coverage_counters[chosen] += 1

    for adj in forced_adjustments:
        d = adj['day_idx']
        p = adj['person']
        type_adj = adj['type']
        if type_adj == 'add': final_schedule[p][d] = "T+"
        elif type_adj == 'remove': final_schedule[p][d] = "L*"

    # Llenado final de francotiradores si sobran días
    fill_log = {}
    for name in roster_df['Nombre']:
        if strategy_key == 'sniper':
            sched = final_schedule[name]
            for d in range(total_days - 2):
                if sched[d] == 'V':
                    if sched[d+1] == 'L': final_schedule[name][d+1] = 'V(R)'
                    if sched[d+2] == 'L': final_schedule[name][d+2] = 'V(R)'
        else:
            current = natural_days_count.get(name, 0)
            needed = 39 - current
            if needed > 0:
                available_idx = [i for i, x in enumerate(final_schedule[name]) if x == 'L']
                available_idx = [x for x in available_idx if not str(final_schedule[name][x]).startswith('T')]
                if len(available_idx) >= needed:
                    fill_idxs = get_clustered_dates(available_idx, needed)
                    for idx in fill_idxs:
                        final_schedule[name][idx] = 'V(R)'

    return final_schedule, adjustments_log, person_coverage_counters, fill_log, natural_days_count

def get_work_days_count(final_schedule):
    counts = {}
    for name, sched in final_schedule.items():
        c = 0
        for s in sched:
            if str(s) == 'T' or str(s).startswith('T*') or str(s) == 'T+': c += 1
        counts[name] = c
    return counts

# --- FUNCIÓN EXCEL REPARADA ---
def create_final_excel(schedule, roster_df, year, requests, fill_log, counters, night_periods, adjustments_log, strategy_key="standard"):
    wb = Workbook()
    
    # Estilos
    s_T = PatternFill("solid", fgColor="C6EFCE"); s_V = PatternFill("solid", fgColor="FFC000") # Oro Fuerte
    s_VR = PatternFill("solid", fgColor="FFFFE0"); s_Cov = PatternFill("solid", fgColor="FFC7CE")
    s_L = PatternFill("solid", fgColor="F2F2F2"); s_Night = PatternFill("solid", fgColor="A6A6A6")
    s_Extra = PatternFill("solid", fgColor="ADD8E6"); s_Free = PatternFill("solid", fgColor="E6E6FA")
    s_VL = PatternFill("solid", fgColor="FFE699") # Oro Suave (Vacaciones en día Libre)
    
    font_bold = Font(bold=True); font_red = Font(color="9C0006", bold=True)
    align_c = Alignment(horizontal="center", vertical="center")
    border_thin = Side(border_style="thin", color="000000")
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # HOJA 1: CUADRANTE
    ws1 = wb.active; ws1.title = "Cuadrante"
    ws1.column_dimensions['A'].width = 20
    for i in range(2, 34): ws1.column_dimensions[get_column_letter(i)].width = 4
    
    curr_row = 1
    for t in TEAMS:
        ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=32)
        cell_title = ws1.cell(curr_row, 1, f"TURNO {t}"); cell_title.font = Font(bold=True, color="FFFFFF"); cell_title.fill = PatternFill("solid", fgColor="000080"); cell_title.alignment = align_c
        curr_row += 2
        members = roster_df[roster_df['Turno'] == t].copy()
        role_order = ["Jefe", "Subjefe", "Conductor", "Bombero"]
        members['sort_key'] = members['Rol'].apply(lambda x: role_order.index(x))
        members = members.sort_values(by=['sort_key', 'Nombre'])
        for _, p in members.iterrows():
            nm = p['Nombre']; role = p['Rol']
            ws1.cell(curr_row, 1, f"{nm} ({role})").font = font_bold
            for d in range(1, 32): c = ws1.cell(curr_row, d+1, d); c.alignment = align_c; c.font = font_bold; c.border = border_all; c.fill = PatternFill("solid", fgColor="E0E0E0")
            curr_row += 1
            for m_idx, mes in enumerate(MESES):
                ws1.cell(curr_row, 1, mes).font = font_bold; ws1.cell(curr_row, 1).border = border_all
                d_month = calendar.monthrange(year, m_idx+1)[1]
                for d in range(1, 32):
                    cell = ws1.cell(curr_row, d+1); cell.border = border_all; cell.alignment = align_c
                    if d <= d_month:
                        dt = datetime.date(year, m_idx+1, d); d_y = dt.timetuple().tm_yday - 1
                        st_val = schedule[nm][d_y]
                        fill = s_L; val = ""
                        
                        # Lógica de pintado mejorada
                        if st_val == 'T': fill = s_T; val = "T"
                        elif st_val == 'V': fill = s_V; val = "V"
                        elif st_val == 'V(L)': fill = s_VL; val = "V" # Se ve como V pero suave
                        elif st_val == 'V(R)': 
                            fill = s_VR; val = "v"
                            if strategy_key == 'sniper': fill = s_V; val = "V" 
                        elif str(st_val).startswith('T*'): 
                            fill = s_Cov; cell.font = font_red
                            raw_name = st_val.split('(')[1][:-1]
                            cov_p = roster_df[roster_df['Nombre'] == raw_name].iloc[0]
                            val = get_short_id(cov_p['Nombre'], cov_p['Rol'], cov_p['Turno'])
                        elif st_val == 'T+': fill = s_Extra; val = "T+"
                        elif st_val == 'L*': fill = s_Free; val = "L"
                        
                        if is_in_night_period(d_y, year, night_periods): fill = s_Night
                        cell.fill = fill; cell.value = val
                    else: cell.fill = PatternFill("solid", fgColor="808080")
                curr_row += 1
            curr_row += 2 
    
    # HOJA 2: ESTADISTICAS
    ws2 = wb.create_sheet("Estadísticas")
    headers = ["Nombre", "Turno", "Puesto", "Días Trabajados", "Gastado (T)", "Coberturas (T*)", "Total Vacs (Nat)"]
    ws2.append(headers)
    
    # Calcular contadores REALES basados en requests (no en grid)
    real_vac_counts = {name: 0 for name in roster_df['Nombre']}
    for req in requests:
        dur = (req['Fin'] - req['Inicio']).days + 1
        real_vac_counts[req['Nombre']] += dur

    for _, p in roster_df.iterrows():
        name = p['Nombre']; sch = schedule[name]
        
        # Créditos T gastados (Contamos 'V' en el grid original antes de coberturas)
        # Como el grid ya tiene T*, es difícil recuperar. Mejor aproximar o recalcular.
        # Simplificación aceptable: Contamos V + T* (si era V)
        # O mejor: contamos créditos desde los requests
        v_credits = 0
        base_sch_turn, _ = generate_base_schedule(year)
        for req in requests:
            if req['Nombre'] == name:
                s = req['Inicio'].timetuple().tm_yday - 1
                e = req['Fin'].timetuple().tm_yday - 1
                for k in range(s, e+1):
                    if base_sch_turn[p['Turno']][k] == 'T': v_credits += 1

        t_cover = counters[name]
        
        # El total de naturales viene directo de la suma de requests (INFALIBLE)
        v_natural = real_vac_counts[name]
        
        total_worked = 0
        for s in sch:
            if str(s) == 'T' or str(s).startswith('T*') or str(s) == 'T+': total_worked += 1
            
        ws2.append([name, p['Turno'], p['Rol'], total_worked, v_credits, t_cover, v_natural])

    # HOJA 3: DETALLE PERIODOS
    ws3 = wb.create_sheet("Listado Periodos")
    ws3.append(["Nombre", "Rol", "Turno", "Inicio", "Fin", "Días"])
    reqs_data = []
    for req in requests:
        p_row = roster_df[roster_df['Nombre'] == req['Nombre']].iloc[0]
        dur = (req['Fin'] - req['Inicio']).days + 1
        reqs_data.append({
            "Nombre": req['Nombre'], "Rol": p_row['Rol'], "Turno": p_row['Turno'],
            "Inicio": req['Inicio'], "Fin": req['Fin'], "Días": dur
        })
    reqs_data.sort(key=lambda x: (x['Turno'], x['Rol'], x['Nombre'], x['Inicio']))
    for r in reqs_data:
        ws3.append([r['Nombre'], r['Rol'], r['Turno'], r['Inicio'], r['Fin'], r['Días']])

    # HOJA 4: AJUSTES
    ws4 = wb.create_sheet("Ajustes")
    ws4.append(["Fecha", "Cubre", "Ausente"])
    for d, c, a in adjustments_log:
        dt = datetime.date(year, 1, 1) + datetime.timedelta(days=d)
        ws4.append([dt.strftime("%d/%m/%Y"), c, a])
    
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ==============================================================================
# 3. INTERFAZ STREAMLIT
# ==============================================================================

if 'raw_requests_df' not in st.session_state:
    st.session_state.raw_requests_df = pd.DataFrame(columns=["Nombre", "Inicio", "Fin"])
if 'forced_adjustments' not in st.session_state:
    st.session_state.forced_adjustments = []
if 'locked_result' not in st.session_state: st.session_state.locked_result = None
if 'nights' not in st.session_state: st.session_state.nights = []
if 'roster_data' not in st.session_state: st.session_state.roster_data = pd.DataFrame(DEFAULT_ROSTER)

current_requests = st.session_state.raw_requests_df.to_dict('records')
year_val = 2026 
stats = calculate_stats(st.session_state.roster_data, current_requests, year_val)

with st.sidebar:
    st.header("Panel de Control")
    def clear_on_strategy_change():
        st.session_state.raw_requests_df = pd.DataFrame(columns=["Nombre", "Inicio", "Fin"])
        st.session_state.locked_result = None
        st.toast("🧹 Estrategia cambiada: Se han reseteado las vacaciones.", icon="⚠️")

    strategy_key = st.selectbox("🎯 Estrategia", list(STRATEGIES.keys()), format_func=lambda x: STRATEGIES[x]['name'], on_change=clear_on_strategy_change)
    
    with st.expander("🌑 Configurar Nocturnas"):
        c1, c2 = st.columns(2)
        d_start = c1.date_input("Inicio", value=None)
        d_end = c2.date_input("Fin", value=None)
        if st.button("Añadir Periodo"):
            if d_start and d_end:
                st.session_state.nights.append((d_start, d_end))
                st.success("Añadido")
        st.write(f"Periodos: {len(st.session_state.nights)}")
        if st.button("Limpiar Nocturnas"): st.session_state.nights = []
        up_n = st.file_uploader("Subir Excel Nocturnas", type=['xlsx'])
        if up_n:
            try:
                df_n = pd.read_excel(up_n)
                for _, row in df_n.iterrows():
                     if not pd.isnull(row.iloc[0]):
                         d1 = pd.to_datetime(row.iloc[0]).date(); d2 = pd.to_datetime(row.iloc[1]).date()
                         st.session_state.nights.append((d1, d2))
                st.success("Cargado")
            except: pass

    st.markdown("---")
    st.subheader("💾 Guardar Progreso")
    if not st.session_state.raw_requests_df.empty:
        csv_buffer = st.session_state.raw_requests_df.to_csv(index=False).encode('utf-8')
        st.download_button("📥 Bajar Backup", data=csv_buffer, file_name="mis_vacaciones.csv", mime="text/csv")
    
    uploaded_file = st.file_uploader("📤 Restaurar Backup", type=["csv"])
    if uploaded_file is not None:
        try:
            df_loaded = pd.read_csv(uploaded_file)
            df_loaded['Inicio'] = pd.to_datetime(df_loaded['Inicio']).dt.date
            df_loaded['Fin'] = pd.to_datetime(df_loaded['Fin']).dt.date
            if st.button("Aplicar Backup"):
                st.session_state.raw_requests_df = df_loaded
                st.session_state.locked_result = None
                st.rerun()
        except Exception as e: st.error(f"Error: {e}")

    st.markdown("---")
    year_val = st.number_input("Año", value=2026)
    
    c_gen_1, c_gen_2 = st.columns(2)
    with c_gen_1:
        if st.button("🎲 Rellenar Automático", type="primary"):
            current_reqs_dict = st.session_state.raw_requests_df.to_dict('records')
            new_reqs = auto_generate_schedule(st.session_state.roster_data, year_val, st.session_state.nights, strategy_key, current_reqs_dict)
            if new_reqs:
                df_new = pd.DataFrame(new_reqs)
                st.session_state.raw_requests_df = pd.concat([st.session_state.raw_requests_df, df_new], ignore_index=True)
                st.session_state.forced_adjustments = []
                st.session_state.locked_result = None 
                st.toast(f"✅ Se han añadido {len(new_reqs)} periodos nuevos.", icon="🤖")
                st.rerun()
            else: st.warning("No se han encontrado huecos para añadir más vacaciones.")
    with c_gen_2:
        if st.button("🗑️ Borrar Todo"):
            st.session_state.raw_requests_df = pd.DataFrame(columns=["Nombre", "Inicio", "Fin"])
            st.session_state.locked_result = None
            st.rerun()

st.divider()
st.subheader("🌍 Ocupación Global")
st.markdown(render_global_occupation_calendar(year_val, st.session_state.roster_data, current_requests, st.session_state.nights), unsafe_allow_html=True)
st.divider()

c_main, c_vis = st.columns([1, 2])
with c_main:
    st.subheader("2. Selección Manual")
    all_names = st.session_state.roster_data['Nombre'].tolist()
    names_sorted = sorted(all_names, key=lambda x: (0 if "Jefe" in x else 1 if "Subjefe" in x else 2 if "Cond" in x else 3))
    selected_person = st.selectbox("Selecciona Trabajador:", names_sorted)
    
    if selected_person:
        st.markdown("---")
        curr_stats = stats.get(selected_person, {'credits': 0, 'natural': 0})
        c = curr_stats['credits']
        remaining = 13 - c
        st.metric("Créditos", f"{c} / 13", delta=remaining)
        
        my_reqs = [r for r in current_requests if r['Nombre'] == selected_person]
        month_range = st.select_slider("📅 Meses:", options=MESES, value=(MESES[0], MESES[-1]))
        options, conflicts_log, occ_map_T, d_absent, d_roles = get_available_blocks_for_person(selected_person, st.session_state.roster_data, current_requests, year_val, st.session_state.nights, month_range, strategy_key)
        
        p_row = st.session_state.roster_data[st.session_state.roster_data['Nombre'] == selected_person].iloc[0]
        p_turn = p_row['Turno']
        base_sch, _ = generate_base_schedule(year_val)
        
        used_counts = {}
        for r in my_reqs:
            dur = (r['Fin'] - r['Inicio']).days + 1
            s_idx = r['Inicio'].timetuple().tm_yday - 1
            e_idx = r['Fin'].timetuple().tm_yday - 1
            cred = 0
            for d_i in range(s_idx, e_idx + 1):
                if base_sch[p_turn][d_i] == 'T': cred += 1
            key = (dur, cred)
            used_counts[key] = used_counts.get(key, 0) + 1

        recipe = STRATEGIES[strategy_key]['auto_recipe']
        limit_counts = {}
        for item in recipe:
            key = (item['dur'], item['target'])
            limit_counts[key] = limit_counts.get(key, 0) + 1
        
        block_defs = STRATEGIES[strategy_key]['blocks']
        tab_labels = [b['label'] for b in block_defs] + ["✨ A Medida (Libre)"]
        tabs = st.tabs(tab_labels)
        
        for i, b_def in enumerate(block_defs):
            with tabs[i]:
                d_val = b_def['dur']; c_val = b_def['cred']
                used = used_counts.get((d_val, c_val), 0)
                limit = limit_counts.get((d_val, c_val), 0)
                
                c_info, c_bar = st.columns([1, 3])
                with c_info: st.metric(label="Usados", value=f"{used} / {limit}")
                with c_bar:
                    if limit > 0: st.progress(min(used/limit, 1.0), text=f"Progreso: {int(min(used/limit, 1.0)*100)}%")
                    else: st.info("Ilimitado")
                st.divider()
                
                with st.expander("🛠️ Crear Combo a Medida (Manual)", expanded=False):
                    st.caption(f"Sumar: **{d_val} días** y **{c_val} créditos**.")
                    c1, c2 = st.columns(2)
                    with c1:
                        d_range_1 = st.date_input("Fechas P1", value=[], key=f"c1_{i}", format="DD/MM/YYYY")
                        p1_cred = 0; p1_dur = 0
                        if len(d_range_1) == 2:
                            s1, e1 = d_range_1; p1_dur = (e1 - s1).days + 1; s1_idx = s1.timetuple().tm_yday - 1
                            for k in range(p1_dur):
                                if base_sch[p_turn][s1_idx + k] == 'T': p1_cred += 1
                            is_val_1, r1, _ = analyze_slot(s1_idx, p1_dur, p_row, occ_map_T, base_sch, year_val, st.session_state.nights, d_absent, d_roles)
                            if not is_val_1: st.error(f"P1: {r1}")
                            else: st.success(f"P1 OK ({p1_dur}d/{p1_cred}c)")
                    with c2:
                        d_range_2 = st.date_input("Fechas P2", value=[], key=f"c2_{i}", format="DD/MM/YYYY")
                        p2_cred = 0; p2_dur = 0
                        if len(d_range_2) == 2:
                            s2, e2 = d_range_2; p2_dur = (e2 - s2).days + 1; s2_idx = s2.timetuple().tm_yday - 1
                            for k in range(p2_dur):
                                if base_sch[p_turn][s2_idx + k] == 'T': p2_cred += 1
                            is_val_2, r2, _ = analyze_slot(s2_idx, p2_dur, p_row, occ_map_T, base_sch, year_val, st.session_state.nights, d_absent, d_roles)
                            if not is_val_2: st.error(f"P2: {r2}")
                            else: st.success(f"P2 OK ({p2_dur}d/{p2_cred}c)")
                    
                    if len(d_range_1) == 2 and len(d_range_2) == 2:
                        total_dur = p1_dur + p2_dur; total_cred = p1_cred + p2_cred
                        st.markdown("---")
                        if total_dur == d_val and total_cred == c_val:
                            if st.button("💾 Guardar Combo", key=f"save_combo_{i}"):
                                new_reqs = [{"Nombre": selected_person, "Inicio": d_range_1[0], "Fin": d_range_1[1]}, {"Nombre": selected_person, "Inicio": d_range_2[0], "Fin": d_range_2[1]}]
                                st.session_state.raw_requests_df = pd.concat([st.session_state.raw_requests_df, pd.DataFrame(new_reqs)], ignore_index=True)
                                st.session_state.locked_result = None; st.rerun()
                        else: st.error(f"Total: {total_dur}d (Busca {d_val}) | {total_cred}c (Busca {c_val})")

                st.divider()
                available_opts = options.get(b_def['label'], [])
                if available_opts:
                    with st.container(height=300):
                        for opt in available_opts[:100]: 
                            btn_key = f"add_{selected_person}_{i}_{opt['start'] if opt['type']=='single' else opt['parts'][0]['start_idx']}"
                            if st.button(f"➕ {opt['label']}", key=btn_key):
                                new_reqs = []
                                if opt.get('type') == 'combo':
                                    part1 = opt['parts'][0]; part2 = opt['parts'][1]
                                    new_reqs = [{"Nombre": selected_person, "Inicio": part1['date_obj'], "Fin": part1['date_obj'] + timedelta(days=part1['dur']-1)},
                                                {"Nombre": selected_person, "Inicio": part2['date_obj'], "Fin": part2['date_obj'] + timedelta(days=part2['dur']-1)}]
                                else:
                                    new_reqs = [{"Nombre": selected_person, "Inicio": opt['start'], "Fin": opt['end']}]
                                st.session_state.raw_requests_df = pd.concat([st.session_state.raw_requests_df, pd.DataFrame(new_reqs)], ignore_index=True)
                                st.session_state.locked_result = None; st.rerun()
                else:
                    st.warning("⚠️ Sin huecos.")
                    conflicts = conflicts_log.get(b_def['label'], [])
                    if conflicts:
                        with st.expander("🕵️ ¿Por qué?"):
                            conflict_summary = {}
                            for c in conflicts: conflict_summary[c['reason']] = conflict_summary.get(c['reason'], 0) + 1
                            for r, count in conflict_summary.items(): st.caption(f"- {r} ({count} días)")

        with tabs[-1]:
            st.info("🛠️ **Selector Libre:** Elige cualquier fecha.")
            d_range_free = st.date_input("Rango", value=[], key="free_range", format="DD/MM/YYYY")
            if len(d_range_free) == 2:
                s_free, e_free = d_range_free
                dur_free = (e_free - s_free).days + 1
                s_idx_free = s_free.timetuple().tm_yday - 1
                cred_free = 0
                for k in range(dur_free):
                    if base_sch[p_turn][s_idx_free + k] == 'T': cred_free += 1
                st.write(f"**Resumen:** {dur_free} días | {cred_free} créditos")
                is_val_free, r_free, _ = analyze_slot(s_idx_free, dur_free, p_row, occ_map_T, base_sch, year_val, st.session_state.nights, d_absent, d_roles)
                if is_val_free:
                    if st.button("💾 Guardar Personalizado", type="primary"):
                        new_req = {"Nombre": selected_person, "Inicio": s_free, "Fin": e_free}
                        st.session_state.raw_requests_df = pd.concat([st.session_state.raw_requests_df, pd.DataFrame([new_req])], ignore_index=True)
                        st.session_state.locked_result = None; st.rerun()
                else: st.error(f"🚫 {r_free}")

        st.markdown("---")
        st.write(f"**Mis Periodos:**")
        if not my_reqs: st.caption("Ninguno")
        else:
            for i, r in enumerate(my_reqs):
                c1, c2 = st.columns([4, 1])
                c1.write(f"{r['Inicio'].strftime('%d/%m')} - {r['Fin'].strftime('%d/%m')}")
                if c2.button("🗑️", key=f"del_{selected_person}_{i}"):
                    df = st.session_state.raw_requests_df
                    idx_to_drop = df[(df['Nombre'] == r['Nombre']) & (df['Inicio'] == r['Inicio'])].index
                    st.session_state.raw_requests_df = df.drop(idx_to_drop)
                    st.session_state.locked_result = None; st.rerun()

with c_vis:
    if selected_person:
        p_row = st.session_state.roster_data[st.session_state.roster_data['Nombre'] == selected_person].iloc[0]
        turn = p_row['Turno']
        st.subheader(f"3. Visor Turno {turn}")
        base_sch, _ = generate_base_schedule(year_val)
        temp_sch = base_sch[turn].copy()
        my_reqs = [r for r in current_requests if r['Nombre'] == selected_person]
        for r in my_reqs:
            s = r['Inicio'].timetuple().tm_yday - 1
            e = r['Fin'].timetuple().tm_yday - 1
            for d in range(s, e+1):
                if temp_sch[d] == 'T': temp_sch[d] = 'V' 
                else: temp_sch[d] = 'V(L)'
        st.markdown(render_annual_calendar(year_val, turn, base_sch, st.session_state.nights, temp_sch), unsafe_allow_html=True)

st.divider()
st.header("⚙️ Resultados Finales")
if st.button("🔄 Calcular Cuadrante", type="primary"):
    with st.spinner("Procesando..."):
        sch, adj, count, fill, nat_counts = validate_and_generate_final(
            st.session_state.roster_data, 
            st.session_state.raw_requests_df.to_dict('records'), 
            year_val, st.session_state.nights, 
            st.session_state.forced_adjustments, strategy_key
        )
        excel_io = create_final_excel(sch, st.session_state.roster_data, year_val, current_requests, fill, count, st.session_state.nights, adj, strategy_key)
        work_days = get_work_days_count(sch)
        st.session_state.locked_result = {"sch": sch, "adj": adj, "work_days": work_days, "excel": excel_io}
    st.success("¡Calculado!")

if st.session_state.locked_result:
    res = st.session_state.locked_result
    st.download_button("📥 Descargar Excel Final", data=res['excel'], file_name=f"Cuadrante_{year_val}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    cols_eq = st.columns(3)
    for i, (name, count) in enumerate(res['work_days'].items()):
        with cols_eq[i % 3]:
            color = "green" if 121 <= count <= 123 else "red"
            st.markdown(f"**{name}**: <span style='color:{color}'>{count}</span>", unsafe_allow_html=True)

st.markdown("---")
st.markdown("<h5 style='text-align: center; color: #888;'>Diseñado por Marcos Esteban Vives</h5>", unsafe_allow_html=True)

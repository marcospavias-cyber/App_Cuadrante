import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from datetime import timedelta
import io
import random
import calendar
from itertools import groupby
from operator import itemgetter

# ==============================================================================
# 1. CONFIGURACIÓN Y CONSTANTES
# ==============================================================================

st.set_page_config(layout="wide", page_title="Gestor V58.0 - Regla C-Noche-B", page_icon="🚒")

# --- ESTILOS VISUALES ---
st.markdown("""
<style>
    .stApp { background-color: #f8f9fa; }
    h1 { color: #d32f2f; }
    div[data-testid="stMetricValue"] { font-size: 1.4rem; font-weight: bold; }
    .stButton>button { width: 100%; border-radius: 5px; }
</style>
""", unsafe_allow_html=True)

st.markdown("<h5 style='text-align: center; color: #888;'>Arquitectura de Precisión - V58.0 (Turno C Noche + Cobertura B)</h5>", unsafe_allow_html=True)
st.title("🚒 Gestor de Cuadrantes: Versión Definitiva")

TEAMS = ['A', 'B', 'C']
MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# --- ESTRATEGIAS ---
STRATEGIES = {
    "standard": {
        "name": "🛡️ Estándar Elástica (Recomendada)",
        "blocks": [
            {"dur": 10, "cred": 4, "label": "📦 Bloque 10d (4 Cr)"},
            {"dur": 10, "cred": 3, "label": "📦 Bloque 10d (3 Cr)"},
            {"dur": 9,  "cred": 3, "label": "📦 Bloque 9d (3 Cr)"},
            {"dur": 4,  "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1,  "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 10, "target": 4}, {"dur": 10, "target": 3}, 
            {"dur": 10, "target": 3}, {"dur": 9,  "target": 3}
        ]
    },
    "safe": {
        "name": "🔢 Matemática Pura",
        "blocks": [
            {"dur": 12, "cred": 4, "label": "📦 Largo 12d (4 Cr)"},
            {"dur": 9,  "cred": 3, "label": "📦 Medio 9d (3 Cr)"},
            {"dur": 6,  "cred": 2, "label": "📦 Corto 6d (2 Cr)"},
            {"dur": 4,  "cred": 1, "label": "🧩 Relleno 4d (1 Cr)"},
            {"dur": 1,  "cred": 1, "label": "🧩 Día Suelto (1 Cr)"}
        ],
        "auto_recipe": [
            {"dur": 12, "target": 4}, {"dur": 12, "target": 4}, 
            {"dur": 9, "target": 3}, {"dur": 6, "target": 2}
        ]
    }
}

DEFAULT_ROSTER = [
    {"ID_Puesto": "Jefe A",       "Nombre": "Jefe A",       "Turno": "A", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe A",    "Nombre": "Subjefe A",    "Turno": "A", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond A",       "Nombre": "Cond A",       "Turno": "A", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero A1",   "Nombre": "Bombero A1",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A2",   "Nombre": "Bombero A2",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero A3",   "Nombre": "Bombero A3",   "Turno": "A", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Jefe B",       "Nombre": "Jefe B",       "Turno": "B", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe B",    "Nombre": "Subjefe B",    "Turno": "B", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond B",       "Nombre": "Cond B",       "Turno": "B", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero B1",   "Nombre": "Bombero B1",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B2",   "Nombre": "Bombero B2",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero B3",   "Nombre": "Bombero B3",   "Turno": "B", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Jefe C",       "Nombre": "Jefe C",       "Turno": "C", "Rol": "Jefe",       "SV": False},
    {"ID_Puesto": "Subjefe C",    "Nombre": "Subjefe C",    "Turno": "C", "Rol": "Subjefe",    "SV": False},
    {"ID_Puesto": "Cond C",       "Nombre": "Cond C",       "Turno": "C", "Rol": "Conductor",  "SV": False},
    {"ID_Puesto": "Bombero C1",   "Nombre": "Bombero C1",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C2",   "Nombre": "Bombero C2",   "Turno": "C", "Rol": "Bombero",    "SV": False},
    {"ID_Puesto": "Bombero C3",   "Nombre": "Bombero C3",   "Turno": "C", "Rol": "Bombero",    "SV": False},
]

# ==============================================================================
# 2. LÓGICA DE NEGOCIO (CORE)
# ==============================================================================

def get_short_id(name, role, turn):
    if role == "Jefe": return f"J{turn}"
    if role == "Subjefe": return f"S{turn}"
    if role == "Conductor": return f"C{turn}"
    if "Bombero" in name:
        parts = name.split()
        if len(parts) > 1:
            suffix = parts[-1]
            if len(suffix) >= 2: return f"B{suffix[-1]}{turn}"
    return f"{name[:3]}{turn}"

@st.cache_data
def generate_base_schedule(year):
    is_leap = calendar.isleap(year)
    total_days = 366 if is_leap else 365
    status = {'A': 0, 'B': 2, 'C': 1} 
    schedule = {team: [] for team in TEAMS}
    for _ in range(total_days):
        for t in TEAMS:
            schedule[t].append('T' if status[t] == 0 else 'L')
            status[t] = (status[t] + 1) % 3
    return schedule, total_days

def is_in_night_period(day_idx, year, night_periods):
    current_date = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
    for start, end in night_periods:
        if start <= current_date <= end: return True
    return False

@st.cache_data
def get_night_transition_dates(night_periods):
    dates = set()
    for start, end in night_periods:
        dates.add(end) 
    return dates

def calculate_stats(roster_df, requests, year):
    base_sch, _ = generate_base_schedule(year)
    stats = {}
    for _, p in roster_df.iterrows():
        stats[p['Nombre']] = {'credits': 0, 'natural': 0}
    for req in requests:
        name = req['Nombre']
        if name not in stats: continue
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        row = roster_df[roster_df['Nombre'] == name].iloc[0]
        nat = (e_idx - s_idx) + 1
        cred = 0
        for d in range(s_idx, e_idx + 1):
            if base_sch[row['Turno']][d] == 'T': cred += 1
        stats[name]['credits'] += cred
        stats[name]['natural'] += nat
    return stats

def get_clustered_dates(available_idxs, needed_count):
    if not available_idxs: return []
    groups = []
    for k, g in groupby(enumerate(available_idxs), lambda ix: ix[0] - ix[1]):
        groups.append(list(map(itemgetter(1), g)))
    groups.sort(key=len, reverse=True)
    selected = []
    for group in groups:
        if len(selected) < needed_count:
            take = min(len(group), needed_count - len(selected))
            selected.extend(group[:take])
        else: break
    return sorted(selected)

def book_slot_gen(start_idx, duration, person, occupation_map):
    for i in range(start_idx, start_idx + duration):
        if i not in occupation_map: occupation_map[i] = []
        occupation_map[i].append(person)

# --- DETECTOR DE CONFLICTOS (MODIFICADO: PERMISO PARA TURNO C) ---
def analyze_slot(start_idx, duration, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles):
    total_days = len(base_sch['A'])
    if start_idx + duration > total_days: return False, "Fuera de año", None
    
    my_start_natural = start_idx
    my_end_natural = start_idx + duration - 1

    for i in range(start_idx, start_idx + duration):
        # 1. REGLA SAGRADA: Cupo >= 2
        if len(daily_absent[i]) >= 2:
            return False, f"Cupo lleno (2 pers) el día {i+1}", "Sistema"

        # 2. REGLA NOCTURNA (MODIFICADA)
        d_obj = datetime.date(year, 1, 1) + timedelta(days=i)
        if d_obj in transition_dates:
            if base_sch[person['Turno']][i] == 'T': 
                # NUEVA REGLA: Si es turno C, se permite. Si es A o B, se bloquea.
                if person['Turno'] != 'C':
                    return False, "Conflicto Nocturna (Solo Turno C permitido)", "Nocturna"
                # Si es turno C, pasa (True)
        
        # 3. REGLA TURNO
        occupants_T = occupation_map_T.get(i, [])
        for occ in occupants_T:
            if occ['Turno'] == person['Turno']: 
                return False, f"Coincide turno T con {occ['Nombre']}", occ['Nombre']
    
    # 4. REGLA CATEGORÍA
    if person['Rol'] != 'Bombero':
        for d_check in range(my_start_natural, my_end_natural + 1):
            people_today = daily_absent[d_check]
            roles_today = daily_roles[d_check]
            for p_name, p_role in zip(people_today, roles_today):
                if p_name != person['Nombre'] and p_role == person['Rol']:
                    return False, f"Coincidencia Categoría con {p_name}", p_name
                        
    return True, "OK", None

# ==============================================================================
# 3. ALGORITMO GENERADOR (ESTRICTO Y DEMOCRÁTICO)
# ==============================================================================

def auto_generate_schedule(roster_df, year, night_periods, strategy_key, current_reqs):
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    
    # --- Mapas de Ocupación ---
    occupation_map_T = {i:[] for i in range(total_days)}
    daily_absent = {i: [] for i in range(total_days)}
    daily_roles = {i: [] for i in range(total_days)}
    
    for req in current_reqs:
        p_row = roster_df[roster_df['Nombre'] == req['Nombre']]
        if p_row.empty: continue
        p_row = p_row.iloc[0]
        s = req['Inicio'].timetuple().tm_yday - 1
        e = req['Fin'].timetuple().tm_yday - 1
        s = max(0, s); e = min(total_days - 1, e)
        for d in range(s, e + 1):
            daily_absent[d].append(req['Nombre'])
            daily_roles[d].append(p_row['Rol'])
            if base_sch[p_row['Turno']][d] == 'T':
                occupation_map_T[d].append(p_row)

    generated_requests = []
    people = roster_df.to_dict('records')
    
    def get_person_stats(p_name, p_turn, extra_reqs):
        all_r = [r for r in current_reqs if r['Nombre'] == p_name] + \
                [r for r in extra_reqs if r['Nombre'] == p_name]
        c = 0; nat = 0; slots = []
        for r in all_r:
            s = r['Inicio'].timetuple().tm_yday - 1
            e = r['Fin'].timetuple().tm_yday - 1
            dur = (e - s) + 1
            nat += dur
            slots.append((s, dur))
            for k in range(s, e+1):
                if 0 <= k < total_days and base_sch[p_turn][k] == 'T': c += 1
        return c, nat, slots

    RECIPE = STRATEGIES[strategy_key]['auto_recipe']
    
    # ==============================================================================
    # RONDA 1: ESTRATEGIA PRINCIPAL (Ahora Turno C puede pillar Noches)
    # ==============================================================================
    random.shuffle(people)
    
    for person in people:
        cred, nat, my_slots = get_person_stats(person['Nombre'], person['Turno'], generated_requests)
        
        current_recipe = RECIPE.copy()
        current_recipe.sort(key=lambda x: x['dur'], reverse=True)
        
        for block in current_recipe:
            duration = block['dur']
            target = block['target']
            
            if cred + target > 13: continue 
            if nat + duration > 39: continue 
            
            valid_starts = []
            for d in range(0, total_days - duration):
                if len(daily_absent[d]) >= 2: continue
                block_broken = False
                for k in range(d, d+duration):
                    if len(daily_absent[k]) >= 2: block_broken = True; break
                if block_broken: continue

                c = sum([1 for k in range(d, d+duration) if base_sch[person['Turno']][k] == 'T'])
                if c == target:
                     is_valid, _, _ = analyze_slot(d, duration, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles)
                     if is_valid: valid_starts.append(d)
            
            if valid_starts:
                start = random.choice(valid_starts)
                overlap = False
                for ms in my_slots:
                    if not (start + duration - 1 < ms[0] or start > ms[0] + ms[1] - 1): 
                        overlap = True; break
                
                if not overlap:
                    book_slot_gen(start, duration, person, occupation_map_T)
                    for k in range(start, start + duration):
                        daily_absent[k].append(person['Nombre'])
                        daily_roles[k].append(person['Rol'])
                    
                    generated_requests.append({
                        "Nombre": person['Nombre'],
                        "Inicio": datetime.date(year, 1, 1) + timedelta(days=start),
                        "Fin": datetime.date(year, 1, 1) + timedelta(days=start+duration-1)
                    })
                    cred += target
                    nat += duration
                    my_slots.append((start, duration))

    # ==============================================================================
    # RONDA 2: EQUILIBRADO EXHAUSTIVO (Créditos)
    # ==============================================================================
    people.sort(key=lambda x: get_person_stats(x['Nombre'], x['Turno'], generated_requests)[0])
    rescue_blocks = [{"dur": 4, "target": 1}, {"dur": 3, "target": 1}, {"dur": 1, "target": 1}]

    for person in people:
        cred, nat, my_slots = get_person_stats(person['Nombre'], person['Turno'], generated_requests)
        if cred >= 13: continue
        
        all_possible_days = list(range(total_days))
        random.shuffle(all_possible_days)

        for r_block in rescue_blocks:
            while cred + r_block['target'] <= 13:
                duration = r_block['dur']
                target = r_block['target']
                if nat + duration > 39: break 
                
                found_slot = False
                for d in all_possible_days:
                    if d + duration > total_days: continue
                    
                    if len(daily_absent[d]) >= 2: continue
                    block_broken = False
                    for k in range(d, d+duration):
                         if len(daily_absent[k]) >= 2: block_broken = True; break
                    if block_broken: continue
                    
                    c = sum([1 for k in range(d, d+duration) if base_sch[person['Turno']][k] == 'T'])
                    if c >= target:
                         is_valid, _, _ = analyze_slot(d, duration, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles)
                         if is_valid:
                             overlap = False
                             for ms in my_slots:
                                 if not (d + duration - 1 < ms[0] or d > ms[0] + ms[1] - 1): overlap = True; break
                             
                             if not overlap:
                                 book_slot_gen(d, duration, person, occupation_map_T)
                                 for k in range(d, d + duration):
                                     daily_absent[k].append(person['Nombre'])
                                     daily_roles[k].append(person['Rol'])
                                 
                                 generated_requests.append({
                                     "Nombre": person['Nombre'],
                                     "Inicio": datetime.date(year, 1, 1) + timedelta(days=d),
                                     "Fin": datetime.date(year, 1, 1) + timedelta(days=d+duration-1)
                                 })
                                 cred += target
                                 nat += duration
                                 my_slots.append((d, duration))
                                 found_slot = True
                                 break 
                
                if not found_slot: break

    # ==============================================================================
    # RONDA 3: RELLENO FINAL EXHAUSTIVO (Naturales)
    # ==============================================================================
    people.sort(key=lambda x: get_person_stats(x['Nombre'], x['Turno'], generated_requests)[1])

    for person in people:
        cred, nat, my_slots = get_person_stats(person['Nombre'], person['Turno'], generated_requests)
        needed = 39 - nat
        
        if needed > 0:
            potential_days = list(range(total_days))
            random.shuffle(potential_days)
            
            for d in potential_days:
                if needed <= 0: break
                if len(daily_absent[d]) >= 2: continue
                
                overlap = False
                for ms in my_slots:
                    if not (d < ms[0] or d > ms[0] + ms[1] - 1): overlap = True; break
                if overlap: continue

                is_working_day = (base_sch[person['Turno']][d] == 'T')
                
                if is_working_day and cred >= 13: continue
                if is_working_day and (cred + 1 > 13): continue
                
                is_valid, _, _ = analyze_slot(d, 1, person, occupation_map_T, base_sch, year, transition_dates, daily_absent, daily_roles)
                
                if is_valid:
                    if is_working_day:
                         book_slot_gen(d, 1, person, occupation_map_T)
                         cred += 1
                    
                    daily_absent[d].append(person['Nombre'])
                    daily_roles[d].append(person['Rol'])
                    
                    generated_requests.append({
                        "Nombre": person['Nombre'],
                        "Inicio": datetime.date(year, 1, 1) + timedelta(days=d),
                        "Fin": datetime.date(year, 1, 1) + timedelta(days=d)
                    })
                    nat += 1
                    needed -= 1
                    my_slots.append((d, 1))

    return generated_requests

# ==============================================================================
# 4. RENDER Y EXCEL
# ==============================================================================

@st.cache_data
def render_global_occupation_calendar(year, roster_df, requests, night_periods):
    base_sch, total_days = generate_base_schedule(year)
    transition_dates = get_night_transition_dates(night_periods)
    occ_map = {d: [] for d in range(total_days)}
    
    for req in requests:
        name = req['Nombre']
        if name not in roster_df['Nombre'].values: continue
        person_row = roster_df[roster_df['Nombre'] == name].iloc[0]
        turn = person_row['Turno']
        s = req['Inicio'].timetuple().tm_yday - 1
        e = req['Fin'].timetuple().tm_yday - 1
        for d in range(s, e+1):
            occ_map[d].append(get_short_id(name, person_row['Rol'], turn))

    html = "<div style='font-family:monospace; font-size:9px; overflow-x: auto;'>"
    html += """
    <div style='display:flex; gap:10px; margin-bottom:10px; font-size:11px; font-weight:bold;'>
        <span style='background:#d4edda; color:#155724; padding:2px 6px; border:1px solid #c3e6cb;'>🟩 DISPONIBLE</span>
        <span style='background:#FFF3CD; color:#856404; padding:2px 6px; border:1px solid #FFEEBA;'>🟧 ÚLTIMA PLAZA</span>
        <span style='background:#F8D7DA; color:#721c24; padding:2px 6px; border:1px solid #F5C6CB;'>🟥 COMPLETO (2/2)</span>
    </div>
    """
    html += "<div style='display:flex; margin-bottom:2px;'><div style='width:35px;'></div>"
    for d in range(1, 32):
        html += f"<div style='width:32px; text-align:center; color:#888;'>{d}</div>"
    html += "</div>"

    for m_idx, mes in enumerate(MESES):
        m_num = m_idx + 1
        days_in_month = calendar.monthrange(year, m_num)[1]
        html += f"<div style='display:flex; margin-bottom:2px;'><div style='width:35px; font-weight:bold; padding-top:8px;'>{mes}</div>"
        for d in range(1, 32):
            if d <= days_in_month:
                dt = datetime.date(year, m_num, d)
                d_idx = dt.timetuple().tm_yday - 1
                occupants = occ_map[d_idx]
                count = len(occupants)
                
                if count == 0: bg = "#d4edda"; txt_col = "#155724"
                elif count == 1: bg = "#FFF3CD"; txt_col = "#856404"
                else: bg = "#F8D7DA"; txt_col = "#721c24" 

                border = "1px solid #fff"
                if dt in transition_dates: border = "2px solid red"
                label = "<br>".join(occupants)
                html += f"<div style='width:32px; height:30px; background-color:{bg}; color:{txt_col}; text-align:center; border:{border}; border-radius:2px; font-size:8px; line-height:9px; display:flex; align-items:center; justify-content:center;'>{label}</div>"
            else:
                html += "<div style='width:32px;'></div>"
        html += "</div>"
    html += "</div>"
    return html

def render_annual_calendar(year, team, base_sch, night_periods, custom_schedule=None):
    html = f"<div style='font-family:monospace; font-size:10px;'>"
    html += """
    <div style='display:flex; gap:10px; margin-bottom:5px; font-size:11px; font-weight:bold;'>
        <span style='background:#d4edda; color:#155724; padding:2px 5px; border:1px solid #c3e6cb;'>T (Guardia)</span>
        <span style='background:#FFC000; color:#000; padding:2px 5px; border:1px solid #DAA520;'>V (Vacaciones)</span>
        <span style='background:#1E7E34; color:white; padding:2px 5px;'>T (Noche)</span>
    </div>
    """
    html += "<div style='display:flex; margin-bottom:2px;'><div style='width:30px;'></div>"
    for d in range(1, 32):
        html += f"<div style='width:20px; text-align:center; color:#888;'>{d}</div>"
    html += "</div>"
    for m_idx, mes in enumerate(MESES):
        m_num = m_idx + 1
        days_in_month = calendar.monthrange(year, m_num)[1]
        html += f"<div style='display:flex; margin-bottom:2px;'><div style='width:30px; font-weight:bold;'>{mes}</div>"
        for d in range(1, 32):
            if d <= days_in_month:
                dt = datetime.date(year, m_num, d)
                d_idx = dt.timetuple().tm_yday - 1
                state = base_sch[team][d_idx]
                final_val = state
                if custom_schedule: final_val = custom_schedule[d_idx]
                
                bg_color = "#eee"; text_color = "#ccc"; border = "1px solid #fff"
                
                if str(final_val).startswith('T'): 
                    bg_color = "#d4edda"; text_color = "#155724"
                    if is_in_night_period(d_idx, year, night_periods):
                        bg_color = "#1E7E34"; text_color = "white"
                elif final_val == 'V':
                    bg_color = "#FFC000"; text_color = "#000"
                
                if dt in get_night_transition_dates(night_periods): border = "2px solid red"
                html += f"<div style='width:20px; background-color:{bg_color}; color:{text_color}; text-align:center; border:{border}; border-radius:2px;'>{state[0]}</div>"
            else:
                html += "<div style='width:20px;'></div>"
        html += "</div>"
    html += "</div>"
    return html

# --- BUSCADOR DE CANDIDATOS (MODIFICADO: REGLA B CUBRE C EN NOCHES) ---
def get_candidates(person_missing, roster_df, day_idx, current_schedule, year, night_periods, unavailable_map, adjustments_log_current_day=None):
    candidates = []
    missing_role = person_missing['Rol']
    missing_turn = person_missing['Turno']
    blocked_turns = set()
    
    if adjustments_log_current_day:
        for coverer_name in adjustments_log_current_day:
            cov_p = roster_df[roster_df['Nombre'] == coverer_name]
            if not cov_p.empty: blocked_turns.add(cov_p.iloc[0]['Turno'])
            
    # Chequeo si es día de Transición Nocturna
    transition_dates = get_night_transition_dates(night_periods)
    current_date = datetime.date(year, 1, 1) + datetime.timedelta(days=day_idx)
    is_transition_day = (current_date in transition_dates)

    turn_exhausted_from_night = None
    if day_idx > 0:
        prev_day_idx = day_idx - 1
        if is_in_night_period(prev_day_idx, year, night_periods):
            base_sch_temp, _ = generate_base_schedule(year)
            for t in TEAMS:
                if base_sch_temp[t][prev_day_idx] == 'T':
                    turn_exhausted_from_night = t; break
                    
    for _, candidate in roster_df.iterrows():
        cand_name = candidate['Nombre']
        cand_turn = candidate['Turno']
        
        # FILTRO SUPREMO: Si falta C en Noche, SOLO B puede cubrir
        if is_transition_day and missing_turn == 'C':
            if cand_turn != 'B': continue
        
        if cand_turn == missing_turn: continue
        if cand_name in unavailable_map[day_idx]: continue
        
        cand_status = current_schedule[cand_name][day_idx]
        if cand_status != 'L': continue 
        
        if cand_turn in blocked_turns: continue
        if turn_exhausted_from_night and cand_turn == turn_exhausted_from_night: continue
        
        is_compatible = False
        cand_role = candidate['Rol']
        if missing_role == "Jefe" and cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Subjefe" and cand_role in ["Jefe", "Subjefe"]: is_compatible = True
        elif missing_role == "Conductor" and (cand_role == "Conductor" or candidate['SV']): is_compatible = True
        elif missing_role == "Bombero" and (cand_role == "Bombero" or candidate['SV']): is_compatible = True
        
        if is_compatible: candidates.append(cand_name)
        
    return candidates

def validate_and_generate_final(roster_df, requests, year, night_periods, strategy_key="standard"):
    base_schedule_turn, total_days = generate_base_schedule(year)
    final_schedule = {} 
    turn_coverage_counters = {'A': 0, 'B': 0, 'C': 0}
    person_coverage_counters = {name: 0 for name in roster_df['Nombre']}
    name_to_turn = {row['Nombre']: row['Turno'] for _, row in roster_df.iterrows()}
    
    for _, row in roster_df.iterrows():
        final_schedule[row['Nombre']] = base_schedule_turn[row['Turno']].copy()

    day_vacations = {i: [] for i in range(total_days)}
    natural_days_count = {name: 0 for name in roster_df['Nombre']}
    unavailable_map = {i: set() for i in range(total_days)}

    for req in requests:
        name = req['Nombre']
        s_idx = req['Inicio'].timetuple().tm_yday - 1
        e_idx = req['Fin'].timetuple().tm_yday - 1
        duration = (e_idx - s_idx) + 1
        natural_days_count[name] += duration
        
        for d in range(s_idx, e_idx + 1):
            unavailable_map[d].add(name)
            if final_schedule[name][d] == 'T':
                day_vacations[d].append(name)
                final_schedule[name][d] = 'V'
            else:
                final_schedule[name][d] = 'V(L)'

    adjustments_log = []
    
    for d in range(total_days):
        absent_people = day_vacations[d]
        if not absent_people: continue
        
        current_day_coverers = []
        absent_people.sort(key=lambda x: 0 if "Jefe" in x or "Subjefe" in x else 1)

        for name_missing in absent_people:
            person_row = roster_df[roster_df['Nombre'] == name_missing].iloc[0]
            candidates = get_candidates(
                person_row, roster_df, d, final_schedule, year, 
                night_periods, unavailable_map, current_day_coverers
            )
            
            if candidates:
                valid = []
                for c in candidates:
                    # En cobertura forzosa nocturna (B cubre C), relajamos regla descanso si es necesario
                    # Pero intentamos mantenerla por defecto
                    prev = final_schedule[c][d-1] if d > 0 else 'L'
                    prev2 = final_schedule[c][d-2] if d > 1 else 'L'
                    worked_prev = prev.startswith('T')
                    worked_prev2 = prev2.startswith('T')
                    if not (worked_prev and worked_prev2): 
                        valid.append(c)
                        
                if valid:
                    valid.sort(key=lambda x: (turn_coverage_counters[name_to_turn[x]], person_coverage_counters[x], random.random()))
                    chosen = valid[0]
                    final_schedule[chosen][d] = f"T*({name_missing})"
                    adjustments_log.append((d, chosen, name_missing))
                    current_day_coverers.append(chosen)
                    unavailable_map[d].add(chosen)
                    turn_coverage_counters[name_to_turn[chosen]] += 1
                    person_coverage_counters[chosen] += 1

    # Relleno de francotiradores/sobrantes original
    fill_log = {}
    for name in roster_df['Nombre']:
        current = natural_days_count.get(name, 0)
        needed = 39 - current
        if needed > 0:
            available_idx = [i for i, x in enumerate(final_schedule[name]) if x == 'L']
            available_idx = [x for x in available_idx if not str(final_schedule[name][x]).startswith('T')]
            if len(available_idx) >= needed:
                fill_idxs = get_clustered_dates(available_idx, needed)
                for idx in fill_idxs:
                    final_schedule[name][idx] = 'V(R)'

    return final_schedule, adjustments_log, person_coverage_counters, fill_log, natural_days_count

def create_final_excel(schedule, roster_df, year, requests, fill_log, counters, night_periods, adjustments_log, strategy_key="standard"):
    wb = Workbook()
    
    s_T = PatternFill("solid", fgColor="C6EFCE"); s_V = PatternFill("solid", fgColor="FFC000") 
    s_VR = PatternFill("solid", fgColor="FFFFE0"); s_Cov = PatternFill("solid", fgColor="FFC7CE")
    s_L = PatternFill("solid", fgColor="F2F2F2"); s_Night = PatternFill("solid", fgColor="A6A6A6")
    s_Extra = PatternFill("solid", fgColor="ADD8E6"); s_Free = PatternFill("solid", fgColor="E6E6FA")
    s_VL = PatternFill("solid", fgColor="FFE699") 
    
    font_bold = Font(bold=True); font_red = Font(color="9C0006", bold=True)
    align_c = Alignment(horizontal="center", vertical="center")
    border_thin = Side(border_style="thin", color="000000")
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # HOJA 1: CUADRANTE
    ws1 = wb.active; ws1.title = "Cuadrante"
    ws1.column_dimensions['A'].width = 20
    for i in range(2, 34): ws1.column_dimensions[get_column_letter(i)].width = 4
    
    curr_row = 1
    for t in TEAMS:
        ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=32)
        cell_title = ws1.cell(curr_row, 1, f"TURNO {t}"); cell_title.font = Font(bold=True, color="FFFFFF"); cell_title.fill = PatternFill("solid", fgColor="000080"); cell_title.alignment = align_c
        curr_row += 2
        members = roster_df[roster_df['Turno'] == t].copy()
        for _, p in members.iterrows():
            nm = p['Nombre']; role = p['Rol']
            ws1.cell(curr_row, 1, f"{nm} ({role})").font = font_bold
            for d in range(1, 32): c = ws1.cell(curr_row, d+1, d); c.alignment = align_c; c.font = font_bold; c.border = border_all; c.fill = PatternFill("solid", fgColor="E0E0E0")
            curr_row += 1
            for m_idx, mes in enumerate(MESES):
                ws1.cell(curr_row, 1, mes).font = font_bold; ws1.cell(curr_row, 1).border = border_all
                d_month = calendar.monthrange(year, m_idx+1)[1]
                for d in range(1, 32):
                    cell = ws1.cell(curr_row, d+1); cell.border = border_all; cell.alignment = align_c
                    if d <= d_month:
                        dt = datetime.date(year, m_idx+1, d); d_y = dt.timetuple().tm_yday - 1
                        st_val = schedule[nm][d_y]
                        fill = s_L; val = ""
                        
                        if st_val == 'T': fill = s_T; val = "T"
                        elif st_val == 'V': fill = s_V; val = "V"
                        elif st_val == 'V(L)': fill = s_VL; val = "V"
                        elif st_val == 'V(R)': 
                            fill = s_VR; val = "v"
                        elif str(st_val).startswith('T*'): 
                            fill = s_Cov; cell.font = font_red
                            raw_name = st_val.split('(')[1][:-1]
                            cov_p = roster_df[roster_df['Nombre'] == raw_name].iloc[0]
                            val = get_short_id(cov_p['Nombre'], cov_p['Rol'], cov_p['Turno'])
                        elif st_val == 'T+': fill = s_Extra; val = "T+"
                        elif st_val == 'L*': fill = s_Free; val = "L"
                        
                        if is_in_night_period(d_y, year, night_periods): fill = s_Night
                        cell.fill = fill; cell.value = val
                    else: cell.fill = PatternFill("solid", fgColor="808080")
                curr_row += 1
            curr_row += 2 
    
    # HOJA 2: ESTADISTICAS
    ws2 = wb.create_sheet("Estadísticas")
    headers = ["Nombre", "Turno", "Puesto", "Días Trabajados", "Gastado (T)", "Coberturas (T*)", "Total Vacs (Nat)"]
    ws2.append(headers)
    
    real_vac_counts = {name: 0 for name in roster_df['Nombre']}
    for req in requests:
        dur = (req['Fin'] - req['Inicio']).days + 1
        real_vac_counts[req['Nombre']] += dur

    for _, p in roster_df.iterrows():
        name = p['Nombre']; sch = schedule[name]
        
        v_credits = 0
        base_sch_turn, _ = generate_base_schedule(year)
        for req in requests:
            if req['Nombre'] == name:
                s = req['Inicio'].timetuple().tm_yday - 1
                e = req['Fin'].timetuple().tm_yday - 1
                for k in range(s, e+1):
                    if base_sch_turn[p['Turno']][k] == 'T': v_credits += 1

        t_cover = counters[name]
        v_natural = real_vac_counts[name]
        
        total_worked = 0
        for s in sch:
            if str(s) == 'T' or str(s).startswith('T*') or str(s) == 'T+': total_worked += 1
            
        ws2.append([name, p['Turno'], p['Rol'], total_worked, v_credits, t_cover, v_natural])

    # HOJA 3: DETALLE PERIODOS
    ws3 = wb.create_sheet("Listado Periodos")
    ws3.append(["Nombre", "Rol", "Turno", "Inicio", "Fin", "Días"])
    reqs_data = []
    for req in requests:
        p_row = roster_df[roster_df['Nombre'] == req['Nombre']].iloc[0]
        dur = (req['Fin'] - req['Inicio']).days + 1
        reqs_data.append({
            "Nombre": req['Nombre'], "Rol": p_row['Rol'], "Turno": p_row['Turno'],
            "Inicio": req['Inicio'], "Fin": req['Fin'], "Días": dur
        })
    reqs_data.sort(key=lambda x: (x['Turno'], x['Rol'], x['Nombre'], x['Inicio']))
    for r in reqs_data:
        ws3.append([r['Nombre'], r['Rol'], r['Turno'], r['Inicio'], r['Fin'], r['Días']])

    # HOJA 4: AJUSTES
    ws4 = wb.create_sheet("Ajustes")
    ws4.append(["Fecha", "Cubre", "Ausente"])
    for d, c, a in adjustments_log:
        dt = datetime.date(year, 1, 1) + datetime.timedelta(days=d)
        ws4.append([dt.strftime("%d/%m/%Y"), c, a])
    
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ==============================================================================
# 5. UI STREAMLIT (CON BACKUPS Y NOCTURNAS)
# ==============================================================================

if 'raw_requests_df' not in st.session_state:
    st.session_state.raw_requests_df = pd.DataFrame(columns=["Nombre", "Inicio", "Fin"])
if 'nights' not in st.session_state: st.session_state.nights = []
if 'roster_data' not in st.session_state: st.session_state.roster_data = pd.DataFrame(DEFAULT_ROSTER)

current_requests = st.session_state.raw_requests_df.to_dict('records')
year_val = 2026 
stats = calculate_stats(st.session_state.roster_data, current_requests, year_val)

with st.sidebar:
    st.header("Panel de Control")
    strategy_key = st.selectbox("🎯 Estrategia", list(STRATEGIES.keys()), format_func=lambda x: STRATEGIES[x]['name'])
    
    # --- NOCTURNAS ---
    with st.expander("🌑 Gestión Nocturnas"):
        c1, c2 = st.columns(2)
        d_start = c1.date_input("Inicio", value=None)
        d_end = c2.date_input("Fin", value=None)
        if st.button("Añadir Periodo"):
            if d_start and d_end:
                st.session_state.nights.append((d_start, d_end))
                st.success("Añadido")
        st.write(f"Periodos activos: {len(st.session_state.nights)}")
        if st.button("Limpiar Nocturnas"): st.session_state.nights = []
        
        up_n = st.file_uploader("Subir Excel Nocturnas", type=['xlsx'], key="night_up")
        if up_n:
            try:
                df_n = pd.read_excel(up_n)
                loaded_c = 0
                for _, row in df_n.iterrows():
                    if not pd.isnull(row.iloc[0]) and not pd.isnull(row.iloc[1]):
                        d1 = pd.to_datetime(row.iloc[0]).date()
                        d2 = pd.to_datetime(row.iloc[1]).date()
                        st.session_state.nights.append((d1, d2))
                        loaded_c += 1
                st.success(f"Cargados {loaded_c} periodos")
            except Exception as e: st.error(f"Error: {e}")

    # --- BACKUP ---
    st.markdown("---")
    st.subheader("💾 Datos y Backup")
    if not st.session_state.raw_requests_df.empty:
        csv_data = st.session_state.raw_requests_df.to_csv(index=False).encode('utf-8')
        st.download_button("📥 Descargar Backup (.csv)", csv_data, "backup_vacaciones.csv", "text/csv")
    
    up_csv = st.file_uploader("📤 Restaurar Backup (.csv)", type=["csv"], key="csv_up")
    if up_csv:
        try:
            df_up = pd.read_csv(up_csv)
            if {"Nombre", "Inicio", "Fin"}.issubset(df_up.columns):
                df_up['Inicio'] = pd.to_datetime(df_up['Inicio']).dt.date
                df_up['Fin'] = pd.to_datetime(df_up['Fin']).dt.date
                if st.button("Aplicar CSV Cargado"):
                    st.session_state.raw_requests_df = df_up
                    st.rerun()
            else: st.error("CSV inválido")
        except Exception as e: st.error(f"Error: {e}")

    # --- ACCIONES PRINCIPALES ---
    st.markdown("---")
    if st.button("🗑️ Resetear Todo", type="secondary"):
        st.session_state.raw_requests_df = pd.DataFrame(columns=["Nombre", "Inicio", "Fin"])
        st.rerun()

    st.markdown("---")
    if st.button("🎲 Rellenar Automático (Estricto)", type="primary"):
        with st.spinner("Optimizando cuadrante (C-Night + B-Cover)..."):
            new_reqs = auto_generate_schedule(st.session_state.roster_data, year_val, st.session_state.nights, strategy_key, current_requests)
            if new_reqs:
                df_new = pd.DataFrame(new_reqs)
                st.session_state.raw_requests_df = pd.concat([st.session_state.raw_requests_df, df_new], ignore_index=True)
                st.success(f"Generados {len(new_reqs)} periodos nuevos.")
                st.rerun()
            else:
                st.warning("No se encontraron más huecos disponibles.")

# MAIN UI
st.subheader("🌍 Mapa de Calor (Ocupación Global)")
st.caption("Verde: Libre | Rojo: Cupo lleno (2 personas)")
st.markdown(render_global_occupation_calendar(year_val, st.session_state.roster_data, current_requests, st.session_state.nights), unsafe_allow_html=True)

col_1, col_2 = st.columns([1, 2])
with col_1:
    st.subheader("Selección Manual")
    sel_person = st.selectbox("Bombero:", st.session_state.roster_data['Nombre'])
    if sel_person:
        curr_s = stats[sel_person]
        st.metric("Créditos", f"{curr_s['credits']} / 13", delta=13-curr_s['credits'])
        st.metric("Naturales", f"{curr_s['natural']} / 39", delta=39-curr_s['natural'])
        d_range = st.date_input("Elegir fechas", [])
        if len(d_range) == 2:
            if st.button("Añadir"):
                new_row = {"Nombre": sel_person, "Inicio": d_range[0], "Fin": d_range[1]}
                st.session_state.raw_requests_df = pd.concat([st.session_state.raw_requests_df, pd.DataFrame([new_row])], ignore_index=True)
                st.rerun()
        my_reqs = st.session_state.raw_requests_df[st.session_state.raw_requests_df['Nombre'] == sel_person]
        if not my_reqs.empty: st.dataframe(my_reqs)

with col_2:
    if sel_person:
        st.subheader(f"Calendario de {sel_person}")
        p_data = st.session_state.roster_data[st.session_state.roster_data['Nombre']==sel_person].iloc[0]
        base_sch, _ = generate_base_schedule(year_val)
        preview_sch = base_sch[p_data['Turno']].copy()
        for r in current_requests:
            if r['Nombre'] == sel_person:
                s = r['Inicio'].timetuple().tm_yday - 1
                e = r['Fin'].timetuple().tm_yday - 1
                for k in range(s, e+1): preview_sch[k] = 'V'
        st.markdown(render_annual_calendar(year_val, p_data['Turno'], base_sch, st.session_state.nights, preview_sch), unsafe_allow_html=True)

st.divider()
if st.button("📥 Generar Excel Final", type="primary"):
    sch, adj, counts, fill_log, nat_counts = validate_and_generate_final(st.session_state.roster_data, current_requests, year_val, st.session_state.nights)
    excel_file = create_final_excel(sch, st.session_state.roster_data, year_val, current_requests, fill_log, counts, st.session_state.nights, adj)
    st.download_button("Descargar Cuadrante .xlsx", excel_file, f"Cuadrante_{year_val}.xlsx")
